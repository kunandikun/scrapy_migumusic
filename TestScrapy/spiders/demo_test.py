import re
import lxml
import html5lib
import requests
import scrapy
from scrapy import Request, Selector
import copy
import openpyxl
from TestScrapy.items import TestscrapyItem

class DemoTestSpider(scrapy.Spider):
    name = 'demo_test'
    start_urls =['https://music.migu.cn/v3']

    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36'}

    res = requests.get(url=start_urls[0], headers=headers)
    res.encoding = res.apparent_encoding
    html_text = res.text

    content = html5lib.parse(html_text, treebuilder="lxml", namespaceHTMLElements=False)
    new_urls = content.xpath('//ul[@class="music show"]/li[4]/a/@href')[0]

    data = openpyxl.Workbook()  # 新建工作簿
    data.create_sheet()
    table = data.active
    col = ['标题1', '标题2', '排名1', '标题3', '排名', '歌名', '作者', '链接']
    for i in range(len(col)):
        table.cell(1, i + 1, col[i])

    data.save('migu_music.xlsx')

    def start_requests(self):
        yield Request(url=self.new_urls,callback=self.parse,dont_filter=True,headers=self.headers)
        yield Request(url=self.new_urls, callback=self.parse1, headers=self.headers, dont_filter=True)

    def parse1(self, response):
        selector = Selector(response)
        alists = selector.xpath('//div[@class="board-sord"]//li/a/@href').extract()
        for a in alists:
            item=TestscrapyItem()
            y_id=''
            a = 'https://music.migu.cn' + a
            if re.search('[^0-9]$', a):
                if re.search('/mv$', a):
                    yield Request(url=a, callback=self.parse_mv, dont_filter=True, headers=self.headers,meta={'item':item,'id':y_id})
                    continue
                if re.search('/newalbum$', a):
                    yield Request(url=a, callback=self.parse_newalbum, dont_filter=True, headers=self.headers,meta={'item':item})
                    continue
                yield Request(url=a, callback=self.parse, dont_filter=True, headers=self.headers)
            else:
                 yield Request(url=a, callback=self.parse_yuebang, dont_filter=True, headers=self.headers)

    def parse_yuebang(self, response):
        selector = Selector(response)
        hrefs=selector.xpath('//div[@class="mb-cover"]/a/@href').extract()
        title_on = selector.xpath('//div[@class="sord-item on"]/div[@class="sord-main"]/text()').extract()
        title_li=selector.xpath('//h3[@class="mb-desc"]/text()').extract()

        title_on.extend(title_li)

        for a in hrefs:
            item = TestscrapyItem()
            item['title']=title_on
            id=''
            yield Request(url=a,callback=self.parse_list,dont_filter=True,headers=self.headers,meta={'item':copy.deepcopy(item),'id':id})

    def parse_newalbum(self,response):
        selector = Selector(response)

        title_on = selector.xpath('//div[@class="sord-item on"]/div[@class="sord-main"]/text()').extract()
        title_li = selector.xpath('//div[@class="top-title"]/text()').extract()
        hrefs=selector.xpath('//div[@class="album-item"]/a/@href').extract()
        id=selector.xpath('//div[starts-with(@class,"triangle")]/text()').extract()

        id=[''.join(i.split()) for i in id]

        title_on.extend(title_li)
        for i in range(len(hrefs)):
            a='https://music.migu.cn'+hrefs[i]
            item = TestscrapyItem()
            item['title'] =title_on
            yield Request(url=a,callback=self.parse_list,dont_filter=True,headers=self.headers,meta={'item':copy.deepcopy(item),'id':id[i]})

    def parse_list(self, response):
        item = response.meta['item']
        selector = Selector(response)

        id = selector.xpath('//div[@class="row J_CopySong"]/div[starts-with(@class,"song-index J_SongIndex")]/span/text()').extract()
        title_h1=selector.xpath('//h1[@class="title"]/text()').extract()
        text = selector.xpath('//a[@class="J-btn-share"]/@data-share').extract()

        title_h1 = [''.join(i.split()) for i in title_h1]

        name = []
        href = []
        author = []
        for t in text:
            t = ''.join(t.split())
            newt = t.replace('"', "'").replace("{\'", '{"').replace("\'}", '"}').replace("\':\'", '":"').replace(
                "\',\'", '","')
            newt = eval(newt)
            name.append(newt['title'])
            href.append('https://music.migu.cn' + newt['linkUrl'])
            author.append(newt['singer'])

        id_new=[]
        y_id=response.meta['id']
        for i in id:
            temp = []
            temp.extend([y_id,i])
            id_new.append(temp)

        if 'id' and 'name' and 'href' and 'author' in item:
            item['id'].extend(id_new)
            item['name'].extend(name)
            item['href'].extend(href)
            item['author'].extend(author)
        else:
            item['id'] = id_new
            item['name'] = name
            item['href'] = href
            item['author'] = author
            item['title'].extend(title_h1)

        next_page = selector.xpath('//a[@class="page-c iconfont cf-next-page"]/@href').extract()
        if next_page:
            next_url = response.urljoin('?page=2')
            yield Request(url=next_url, callback=self.parse_list, dont_filter=True, headers=self.headers,
                          meta={'item': copy.deepcopy(item),'id':y_id})
        else:
            yield item

    def parse_mv(self, response):
        item=response.meta['item']
        selector = Selector(response)

        title_on = selector.xpath('//div[@class="sord-item on"]/div[@class="sord-main"]/text()').extract()
        title_li = selector.xpath('//div[@class="top-title"]/text()').extract()
        name=selector.xpath('//div[@class="video-name"]/a/text()').extract()
        href=selector.xpath('//div[@class="video-name"]/a/@href').extract()
        id=selector.xpath('//div[starts-with(@class,"triangle")]/text()').extract()

        author = []
        text= selector.xpath('//div[@class="video-desc"]/comment()').extract()
        for t in text:
            t = re.split('--', str(t))
            t = t[1:-1]
            elem =lxml.etree.HTML(t[0])
            n=elem.xpath('//a/text()')
            if len(n)>1:
                newn=n[0]
                for i in range(1,len(n)):
                    newn=newn+','+n[i]
                author.append(newn)
            else:
                author.extend(n)

        href=['https://music.migu.cn'+a for a in href]

        title_on.extend(title_li)

        id_new = []
        y_id = response.meta['id']
        for i in id:
            temp = []
            temp.extend([y_id, i])
            id_new.append(temp)

        item['name']=name
        item['href']=href
        item['author']=author
        item['title']=title_on
        item['id']=id_new

        yield item

    def parse(self, response):
        item=TestscrapyItem()
        selector = Selector(response)

        id=selector.xpath('//div[starts-with(@class,"song-number")]/text()').extract()
        title_on=selector.xpath('//div[@class="sord-item on"]/div[@class="sord-main"]/text()').extract()
        title_li=selector.xpath('//div[@class="top-title"]/text()').extract()
        text=selector.xpath('//a[@class="J-btn-share"]/@data-share').extract()

        name = []
        href = []
        author = []
        for t in text:
            t = ''.join(t.split())
            newt = t.replace('"', "'").replace("{\'", '{"').replace("\'}", '"}').replace("\':\'", '":"').replace(
                "\',\'", '","')
            newt = eval(newt)
            name.append(newt['title'])
            href.append('https://music.migu.cn' + newt['linkUrl'])
            author.append(newt['singer'])

        id_new = []
        for i in id:
            temp = []
            temp.extend(['', i])
            id_new.append(temp)

        item['id']=id_new
        item['name']=name
        item['href'] = href
        item['author'] = author

        title_on.extend(title_li)
        item['title']=title_on

        yield item










