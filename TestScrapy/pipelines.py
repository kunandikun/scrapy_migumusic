# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import openpyxl
from openpyxl.styles import *
from openpyxl.utils import *

class TestscrapyPipeline:
    def process_item(self, item, spider):
        print(item)
        align=Alignment(horizontal='center',vertical='center',wrap_text=True)
        if len(item['name'])!=0:
            data = openpyxl.load_workbook('migu_music.xlsx')
            table = data.active
            nrows = table.max_row

            table.merge_cells(start_row=nrows + 1, start_column=1, end_row=len(item['name']) + nrows, end_column=1)
            table.merge_cells(start_row=nrows +1 , start_column=2, end_row=len(item['name']) + nrows, end_column=2)
            table.merge_cells(start_row=nrows + 1, start_column=3, end_row=len(item['name']) + nrows, end_column=3)
            table.merge_cells(start_row=nrows + 1, start_column=4, end_row=len(item['name']) + nrows, end_column=4)

            for i in range(len(item['name'])):
                top_left_cell_id1 = table.cell(nrows+1, 3)
                top_left_cell_id1.value = item['id'][i][0]
                top_left_cell_id1.alignment = align

                table_id=table.cell(i + 1 + nrows, 5, item['id'][i][1])
                table_id.alignment=align
                table_name=table.cell(i + 1 + nrows, 6,item['name'][i])
                table_name.alignment=align
                table_author=table.cell(i + 1 + nrows, 7, item['author'][i])
                table_author.alignment=align
                table_href=table.cell(i + 1 + nrows, 8, item['href'][i])
                table_href.alignment=align

                table.row_dimensions[i+1+nrows].height=28
                table.column_dimensions[get_column_letter(6)].width=40
                table.column_dimensions[get_column_letter(7)].width = 50
                table.column_dimensions[get_column_letter(8)].width = 50


            top_left_cell_title1 = table.cell(nrows + 1, 1)
            top_left_cell_title1.value = item['title'][0]
            top_left_cell_title1.alignment=align

            top_left_cell_title2 = table.cell(nrows + 1, 2)
            top_left_cell_title2.value = item['title'][1]
            top_left_cell_title2.alignment = align

            if len(item['title'])==3:
                top_left_cell_title2 = table.cell(nrows + 1, 4)
                top_left_cell_title2.value = item['title'][2]
                top_left_cell_title2.alignment = align

            data.save('migu_music.xlsx')

        return item
